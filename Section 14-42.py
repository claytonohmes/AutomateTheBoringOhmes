import openpyxl
import os

print(os.getcwd())

workbook = openpyxl.load_workbook('example.xlsx')
print(type(workbook))

sheet = workbook['Sheet1']
print(type(sheet))

print(workbook.sheetnames)

print(sheet['A1'])

cell = sheet['A1']

print(cell.value)

print(str(sheet['A1'].value))
print(sheet['B1'].value)
print(sheet['C1'].value)

print(sheet.cell(row=1,column=2))

for i in range(1, 8):
    print(i, sheet.cell(row=i, column=2).value)