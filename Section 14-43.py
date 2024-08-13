import openpyxl, os
wb = openpyxl.Workbook()

print(wb)
sheet = wb['Sheet']

#this is a totally blank new worksheet so the value method is going to return none.
print(sheet['A1'].value == None)

sheet['A1'] = 42
sheet['A2'] = 'Hello'
os.chdir('c:\\users\\cohme\\Desktop')
wb.save('example2.xlsx')

sheet2 = wb.create_sheet()
print(wb.sheetnames)

print(sheet2.title)
sheet2.title = 'My New Sheet Name'
print(sheet2.title)
print(wb.sheetnames)
wb.create_sheet(index=0, title='My Other Sheet')
print(wb.sheetnames)
wb.save('example3.xlsx')